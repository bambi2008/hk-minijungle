from __future__ import annotations

import csv
import math
import shutil
import zipfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
PKG = ROOT / "AquaClick_V0.1C"
OUT = ROOT / "output"
HANDOFF = Path(r"C:\Users\ss\Downloads\AquaClick_Codex_Project_Handoff_2026-07-16.md")


@dataclass(frozen=True)
class HopperConfig:
    name: str
    outer_diameter_mm: float
    top_z_mm: float = 260.0
    wall_mm: float = 2.2
    outlet_x_mm: float = 38.0
    outlet_y_mm: float = -18.0
    outlet_radius_mm: float = 28.0
    floor_low_z_mm: float = 16.0
    floor_high_z_mm: float = 78.0
    safety_headspace_mm: float = 18.0


ROTOR_VARIANTS = {
    "A": {"diameter": 56.0, "height": 26.0, "vanes": 6, "clearance": 1.2},
    "B": {"diameter": 62.0, "height": 28.0, "vanes": 5, "clearance": 1.4},
    "C": {"diameter": 68.0, "height": 30.0, "vanes": 4, "clearance": 1.6},
}

WINDOW_VARIANTS = {
    "A": {"w": 19.5, "h": 20.0},
    "B": {"w": 22.0, "h": 22.0},
    "C": {"w": 26.0, "h": 24.0},
}

CHANNEL_VARIANTS = {
    "standard": {"inner_w": 24.0, "inner_h": 22.0, "length": 82.0, "wall": 2.0, "draft_deg": 2.0, "radius": 3.0},
    "wide": {"inner_w": 30.0, "inner_h": 24.0, "length": 88.0, "wall": 2.2, "draft_deg": 2.0, "radius": 3.5},
}


def reset_dirs() -> None:
    for path in [
        PKG / "cad" / "stl",
        PKG / "drawings",
        PKG / "docs",
        PKG / "data",
        PKG / "reports",
    ]:
        path.mkdir(parents=True, exist_ok=True)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def ascii_stl(path: Path, name: str, triangles: list[tuple[np.ndarray, np.ndarray, np.ndarray]]) -> None:
    def normal(a: np.ndarray, b: np.ndarray, c: np.ndarray) -> np.ndarray:
        n = np.cross(b - a, c - a)
        mag = np.linalg.norm(n)
        return n / mag if mag else np.array([0.0, 0.0, 0.0])

    with path.open("w", encoding="ascii") as f:
        f.write(f"solid {name}\n")
        for a, b, c in triangles:
            n = normal(a, b, c)
            f.write(f"  facet normal {n[0]:.6g} {n[1]:.6g} {n[2]:.6g}\n")
            f.write("    outer loop\n")
            for p in (a, b, c):
                f.write(f"      vertex {p[0]:.6g} {p[1]:.6g} {p[2]:.6g}\n")
            f.write("    endloop\n")
            f.write("  endfacet\n")
        f.write(f"endsolid {name}\n")


def hopper_floor_z(cfg: HopperConfig, x: float, y: float) -> float:
    r = cfg.outer_diameter_mm / 2 - cfg.wall_mm
    dx = x - cfg.outlet_x_mm
    dy = y - cfg.outlet_y_mm
    d = math.sqrt(dx * dx + dy * dy)
    normalized = min(d / (r + cfg.outlet_radius_mm), 1.0)
    return cfg.floor_low_z_mm + (cfg.floor_high_z_mm - cfg.floor_low_z_mm) * normalized ** 1.18


def make_hopper_stl(cfg: HopperConfig) -> dict[str, float]:
    r_outer = cfg.outer_diameter_mm / 2
    r_inner = r_outer - cfg.wall_mm
    seg = 96
    rings = 18
    triangles: list[tuple[np.ndarray, np.ndarray, np.ndarray]] = []
    inner_points: list[list[np.ndarray]] = []

    for i in range(rings + 1):
        rr = r_inner * i / rings
        ring: list[np.ndarray] = []
        for j in range(seg):
            th = 2 * math.pi * j / seg
            x, y = rr * math.cos(th), rr * math.sin(th)
            ring.append(np.array([x, y, hopper_floor_z(cfg, x, y)]))
        inner_points.append(ring)

    for i in range(rings):
        for j in range(seg):
            a = inner_points[i][j]
            b = inner_points[i + 1][j]
            c = inner_points[i + 1][(j + 1) % seg]
            d = inner_points[i][(j + 1) % seg]
            triangles += [(a, b, c), (a, c, d)]

    for j in range(seg):
        th0 = 2 * math.pi * j / seg
        th1 = 2 * math.pi * ((j + 1) % seg) / seg
        inner0 = np.array([r_inner * math.cos(th0), r_inner * math.sin(th0), hopper_floor_z(cfg, r_inner * math.cos(th0), r_inner * math.sin(th0))])
        inner1 = np.array([r_inner * math.cos(th1), r_inner * math.sin(th1), hopper_floor_z(cfg, r_inner * math.cos(th1), r_inner * math.sin(th1))])
        outer0b = np.array([r_outer * math.cos(th0), r_outer * math.sin(th0), inner0[2] - 1.5])
        outer1b = np.array([r_outer * math.cos(th1), r_outer * math.sin(th1), inner1[2] - 1.5])
        outer0t = np.array([r_outer * math.cos(th0), r_outer * math.sin(th0), cfg.top_z_mm])
        outer1t = np.array([r_outer * math.cos(th1), r_outer * math.sin(th1), cfg.top_z_mm])
        inner0t = np.array([r_inner * math.cos(th0), r_inner * math.sin(th0), cfg.top_z_mm])
        inner1t = np.array([r_inner * math.cos(th1), r_inner * math.sin(th1), cfg.top_z_mm])
        triangles += [(outer0b, outer1b, outer1t), (outer0b, outer1t, outer0t)]
        triangles += [(inner1, inner0, inner0t), (inner1, inner0t, inner1t)]
        triangles += [(inner0t, outer0t, outer1t), (inner0t, outer1t, inner1t)]

    path = PKG / "cad" / "stl" / f"aquaclick_v01c_{cfg.name}_eccentric_hopper.stl"
    ascii_stl(path, f"aquaclick_v01c_{cfg.name}_eccentric_hopper", triangles)

    floor_samples = []
    for j in range(seg):
        th = 2 * math.pi * j / seg
        x, y = r_inner * math.cos(th), r_inner * math.sin(th)
        dz = hopper_floor_z(cfg, x, y) - cfg.floor_low_z_mm
        dist = math.sqrt((x - cfg.outlet_x_mm) ** 2 + (y - cfg.outlet_y_mm) ** 2)
        floor_samples.append(math.degrees(math.atan2(dz, max(dist, 1.0))))

    fill_z = cfg.top_z_mm - cfg.safety_headspace_mm
    gross_volume_l = math.pi * r_inner * r_inner * (fill_z - cfg.floor_high_z_mm * 0.55) / 1_000_000
    displaced_l = 0.23 if cfg.outer_diameter_mm <= 180 else 0.25
    usable_l = gross_volume_l - displaced_l
    return {
        "gross_volume_l": gross_volume_l,
        "usable_volume_l": usable_l,
        "fill_z_mm": fill_z,
        "min_floor_angle_deg": min(floor_samples),
        "max_floor_angle_deg": max(floor_samples),
        "stl": path.name,
    }


def make_box_stl(path: Path, name: str, x: float, y: float, z: float) -> None:
    hx, hy, hz = x / 2, y / 2, z / 2
    pts = [np.array(p) for p in [
        (-hx, -hy, -hz), (hx, -hy, -hz), (hx, hy, -hz), (-hx, hy, -hz),
        (-hx, -hy, hz), (hx, -hy, hz), (hx, hy, hz), (-hx, hy, hz),
    ]]
    faces = [(0, 1, 2), (0, 2, 3), (4, 6, 5), (4, 7, 6), (0, 4, 5), (0, 5, 1), (1, 5, 6), (1, 6, 2), (2, 6, 7), (2, 7, 3), (3, 7, 4), (3, 4, 0)]
    ascii_stl(path, name, [(pts[a], pts[b], pts[c]) for a, b, c in faces])


def make_cylinder_stl(path: Path, name: str, diameter: float, height: float, seg: int = 72) -> None:
    r = diameter / 2
    z0, z1 = -height / 2, height / 2
    triangles = []
    center0 = np.array([0.0, 0.0, z0])
    center1 = np.array([0.0, 0.0, z1])
    for j in range(seg):
        th0 = 2 * math.pi * j / seg
        th1 = 2 * math.pi * ((j + 1) % seg) / seg
        p0 = np.array([r * math.cos(th0), r * math.sin(th0), z0])
        p1 = np.array([r * math.cos(th1), r * math.sin(th1), z0])
        p2 = np.array([r * math.cos(th1), r * math.sin(th1), z1])
        p3 = np.array([r * math.cos(th0), r * math.sin(th0), z1])
        triangles += [(p0, p1, p2), (p0, p2, p3), (center0, p1, p0), (center1, p3, p2)]
    ascii_stl(path, name, triangles)


def write_svg(path: Path, title: str, body: str) -> None:
    write_text(path, f"""
<svg xmlns="http://www.w3.org/2000/svg" width="1400" height="900" viewBox="0 0 1400 900">
  <rect width="1400" height="900" fill="#fbfaf7"/>
  <text x="60" y="70" font-family="Arial" font-size="34" font-weight="700">{title}</text>
  {body}
</svg>
""")


def dim_line(x1: int, y1: int, x2: int, y2: int, label: str, tx: int, ty: int) -> str:
    return f"""
  <line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" stroke="#1f2933" stroke-width="2"/>
  <path d="M{x1} {y1} l10 -6 l0 12 z" fill="#1f2933"/>
  <path d="M{x2} {y2} l-10 -6 l0 12 z" fill="#1f2933"/>
  <text x="{tx}" y="{ty}" font-family="Arial" font-size="22">{label}</text>
"""


def make_drawings() -> None:
    write_svg(
        PKG / "drawings" / "AquaClick_V0.1C_Eccentric_Hopper_Section.svg",
        "AquaClick V0.1C - eccentric compound hopper section",
        """
  <path d="M310 150 L1090 150 L970 650 Q700 760 430 650 Z" fill="#e9f3f0" stroke="#1e3d3a" stroke-width="4"/>
  <path d="M430 650 Q775 700 910 470 Q945 405 990 360" fill="none" stroke="#197a75" stroke-width="8"/>
  <circle cx="970" cy="520" r="46" fill="#fff" stroke="#c44736" stroke-width="5"/>
  <text x="120" y="210" font-family="Arial" font-size="24">OD 180 / 185 mm variants</text>
  <text x="120" y="250" font-family="Arial" font-size="24">Offset outlet center: X +38 mm, Y -18 mm</text>
  <text x="120" y="290" font-family="Arial" font-size="24">Lowest floor: Z 16 mm; high floor: Z 78 mm</text>
  <text x="120" y="330" font-family="Arial" font-size="24">Digital check only: real kibble flow still requires EVT</text>
  <line x1="970" y1="520" x2="1180" y2="520" stroke="#c44736" stroke-width="3"/>
  <text x="1190" y="528" font-family="Arial" font-size="22">rotor inlet</text>
""",
    )
    write_svg(
        PKG / "drawings" / "AquaClick_V0.1C_Washable_Chute_Liner.svg",
        "AquaClick V0.1C - washable chute liner concept",
        """
  <rect x="340" y="260" width="560" height="230" rx="12" fill="#eef2f7" stroke="#2d3b52" stroke-width="4"/>
  <rect x="385" y="305" width="470" height="140" rx="8" fill="#fff" stroke="#197a75" stroke-width="4"/>
  <rect x="280" y="330" width="70" height="90" rx="10" fill="#ffe8c7" stroke="#b16a00" stroke-width="4"/>
  <rect x="890" y="330" width="70" height="90" rx="10" fill="#ffe8c7" stroke="#b16a00" stroke-width="4"/>
  <text x="180" y="150" font-family="Arial" font-size="24">Wall: 2.0-2.2 mm | Draft: 2 deg | Corner radius: 3.0-3.5 mm</text>
  <text x="180" y="190" font-family="Arial" font-size="24">Snap tabs are geometry placeholders for OEM DFM review</text>
  <text x="180" y="230" font-family="Arial" font-size="24">Hand-wash claim only: dishwasher not claimed in V0.1C</text>
""",
    )
    write_svg(
        PKG / "drawings" / "AquaClick_V0.1C_Assembly_Boundary.svg",
        "AquaClick V0.1C - washable / non-washable boundary",
        """
  <rect x="130" y="170" width="500" height="560" rx="20" fill="#dff3ed" stroke="#197a75" stroke-width="5"/>
  <rect x="770" y="170" width="500" height="560" rx="20" fill="#f1eef7" stroke="#4c3d72" stroke-width="5"/>
  <text x="220" y="240" font-family="Arial" font-size="30" font-weight="700">Washable zone</text>
  <text x="855" y="240" font-family="Arial" font-size="30" font-weight="700">Dry electrical zone</text>
  <text x="190" y="320" font-family="Arial" font-size="24">food hopper</text>
  <text x="190" y="370" font-family="Arial" font-size="24">rotor removable part</text>
  <text x="190" y="420" font-family="Arial" font-size="24">chute liner</text>
  <text x="190" y="470" font-family="Arial" font-size="24">bowl tray</text>
  <text x="840" y="320" font-family="Arial" font-size="24">motor and gearbox</text>
  <text x="840" y="370" font-family="Arial" font-size="24">main PCB and power</text>
  <text x="840" y="420" font-family="Arial" font-size="24">camera / AI module</text>
  <text x="840" y="470" font-family="Arial" font-size="24">load-cell wiring</text>
  <line x1="635" y1="450" x2="765" y2="450" stroke="#c44736" stroke-width="8"/>
  <text x="610" y="505" font-family="Arial" font-size="24">seal + mechanical lock boundary</text>
""",
    )
    write_svg(
        PKG / "drawings" / "AquaClick_V0.2_GA_Wet_Dry_Section.svg",
        "AquaClick V0.2 - general arrangement wet/dry section",
        f"""
  <rect x="250" y="110" width="500" height="470" rx="18" fill="#dff3ed" stroke="#197a75" stroke-width="5"/>
  <path d="M300 160 L700 160 L650 470 Q500 550 350 470 Z" fill="#f8fffd" stroke="#197a75" stroke-width="4"/>
  <circle cx="620" cy="470" r="42" fill="#fff" stroke="#c44736" stroke-width="5"/>
  <rect x="375" y="585" width="360" height="110" rx="18" fill="#f3f0fa" stroke="#4c3d72" stroke-width="5"/>
  <rect x="805" y="250" width="260" height="190" rx="16" fill="#eef2f7" stroke="#2d3b52" stroke-width="4"/>
  <rect x="805" y="500" width="260" height="120" rx="16" fill="#fff4d6" stroke="#b16a00" stroke-width="4"/>
  <line x1="250" y1="575" x2="1090" y2="575" stroke="#c44736" stroke-width="7" stroke-dasharray="16 10"/>
  <text x="270" y="95" font-family="Arial" font-size="24" font-weight="700">WET ZONE: no PCB / no motor / no wiring</text>
  <text x="385" y="745" font-family="Arial" font-size="24" font-weight="700">DRY BASE: motor + PCB + power</text>
  <text x="820" y="235" font-family="Arial" font-size="22">Click smart module bay</text>
  <text x="820" y="490" font-family="Arial" font-size="22">bowl / load path</text>
  <text x="780" y="570" font-family="Arial" font-size="22" fill="#c44736">wash boundary</text>
  {dim_line(250, 815, 750, 815, "hopper OD target: 180 / 185 mm", 365, 800)}
  {dim_line(1130, 110, 1130, 695, "overall height target <= 305 mm", 1160, 405)}
  <text x="80" y="835" font-family="Arial" font-size="18">Engineering communication drawing only. Not mold data or production drawing.</text>
""",
    )
    write_svg(
        PKG / "drawings" / "AquaClick_V0.2_Washable_Food_Path_Detail.svg",
        "AquaClick V0.2 - washable food path detail",
        f"""
  <path d="M250 150 L820 150 L760 410 Q560 520 330 430 Z" fill="#e9f8f4" stroke="#197a75" stroke-width="5"/>
  <path d="M330 430 Q610 500 720 330" fill="none" stroke="#0f766e" stroke-width="8"/>
  <circle cx="730" cy="418" r="58" fill="#fff" stroke="#c44736" stroke-width="5"/>
  <rect x="760" y="396" width="330" height="86" rx="12" fill="#eef2f7" stroke="#2d3b52" stroke-width="4"/>
  <rect x="800" y="418" width="250" height="42" rx="7" fill="#fff" stroke="#197a75" stroke-width="3"/>
  <rect x="905" y="500" width="210" height="92" rx="18" fill="#fff4d6" stroke="#b16a00" stroke-width="4"/>
  <text x="260" y="120" font-family="Arial" font-size="22">removable hopper</text>
  <text x="620" y="350" font-family="Arial" font-size="22">removable rotor</text>
  <text x="815" y="390" font-family="Arial" font-size="22">washable liner</text>
  <text x="915" y="625" font-family="Arial" font-size="22">removable bowl/tray</text>
  {dim_line(800, 700, 1050, 700, "liner inner width: 24 / 30 mm", 805, 685)}
  {dim_line(1130, 418, 1130, 460, "inner height: 22 / 24 mm", 1160, 447)}
  <text x="90" y="765" font-family="Arial" font-size="21">Design rule: food-contact path comes out as a passive washable assembly.</text>
  <text x="90" y="805" font-family="Arial" font-size="21">No dishwasher claim yet; V1 claim remains rinse / soak / hand-wash unless tested.</text>
""",
    )
    write_svg(
        PKG / "drawings" / "AquaClick_V0.2_Click_Interface_Detail.svg",
        "AquaClick V0.2 - Click interface detail",
        f"""
  <rect x="210" y="170" width="380" height="330" rx="18" fill="#dff3ed" stroke="#197a75" stroke-width="5"/>
  <rect x="660" y="170" width="380" height="330" rx="18" fill="#f3f0fa" stroke="#4c3d72" stroke-width="5"/>
  <circle cx="565" cy="260" r="24" fill="#111827"/>
  <circle cx="685" cy="260" r="24" fill="#111827"/>
  <path d="M570 395 L640 395 L660 430 L590 430 Z" fill="#ffe8c7" stroke="#b16a00" stroke-width="4"/>
  <path d="M660 395 L730 395 L710 430 L640 430 Z" fill="#ffe8c7" stroke="#b16a00" stroke-width="4"/>
  <rect x="580" y="515" width="240" height="42" rx="12" fill="#fff" stroke="#c44736" stroke-width="4"/>
  <line x1="615" y1="170" x2="615" y2="557" stroke="#c44736" stroke-width="5" stroke-dasharray="12 8"/>
  <text x="270" y="135" font-family="Arial" font-size="24" font-weight="700">wet assembly side</text>
  <text x="725" y="135" font-family="Arial" font-size="24" font-weight="700">dry base side</text>
  <text x="520" y="230" font-family="Arial" font-size="22">magnet alignment</text>
  <text x="500" y="382" font-family="Arial" font-size="22">mechanical latch faces</text>
  <text x="835" y="550" font-family="Arial" font-size="22">gasket / labyrinth boundary</text>
  <text x="150" y="650" font-family="Arial" font-size="22">Layer 1: magnets locate and create premium feel.</text>
  <text x="150" y="690" font-family="Arial" font-size="22">Layer 2: mechanical lock handles retention and pet/child misuse.</text>
  <text x="150" y="730" font-family="Arial" font-size="22">Layer 3: gasket/labyrinth protects dry electronics from rinse water and food dust.</text>
  {dim_line(210, 805, 1040, 805, "interface envelope draft: width TBD after teardown", 445, 790)}
""",
    )
    write_svg(
        PKG / "drawings" / "AquaClick_V0.2_Dry_Base_Envelope.svg",
        "AquaClick V0.2 - dry base envelope",
        f"""
  <rect x="230" y="180" width="820" height="460" rx="28" fill="#f3f0fa" stroke="#4c3d72" stroke-width="5"/>
  <circle cx="470" cy="360" r="70" fill="#fff" stroke="#4c3d72" stroke-width="4"/>
  <rect x="620" y="285" width="250" height="150" rx="12" fill="#eef2f7" stroke="#2d3b52" stroke-width="4"/>
  <rect x="390" y="520" width="430" height="65" rx="16" fill="#fff4d6" stroke="#b16a00" stroke-width="4"/>
  <rect x="910" y="265" width="80" height="210" rx="12" fill="#dff3ed" stroke="#197a75" stroke-width="4"/>
  <text x="410" y="360" font-family="Arial" font-size="22">motor</text>
  <text x="680" y="360" font-family="Arial" font-size="22">PCB / power</text>
  <text x="470" y="560" font-family="Arial" font-size="22">load-cell protected path</text>
  <text x="885" y="500" font-family="Arial" font-size="22">module bay / contacts</text>
  <text x="145" y="110" font-family="Arial" font-size="24" font-weight="700">Dry base must survive food dust nearby, but is not washable.</text>
  {dim_line(230, 720, 1050, 720, "base footprint TBD after competitor teardown", 440, 705)}
  <text x="125" y="790" font-family="Arial" font-size="21">No wet-zone electronics. Smart modules remain dry or independently sealed.</text>
""",
    )
    write_svg(
        PKG / "drawings" / "AquaClick_V0.2_Dual_Prototype_Roadmap.svg",
        "AquaClick V0.2 - dual prototype roadmap",
        """
  <rect x="90" y="180" width="530" height="470" rx="20" fill="#e9f8f4" stroke="#197a75" stroke-width="5"/>
  <rect x="780" y="180" width="530" height="470" rx="20" fill="#f3f0fa" stroke="#4c3d72" stroke-width="5"/>
  <text x="145" y="245" font-family="Arial" font-size="30" font-weight="700">Prototype A</text>
  <text x="145" y="285" font-family="Arial" font-size="24">Engineering feeder mule</text>
  <text x="830" y="245" font-family="Arial" font-size="30" font-weight="700">Prototype B</text>
  <text x="830" y="285" font-family="Arial" font-size="24">Click experience model</text>
  <text x="145" y="360" font-family="Arial" font-size="22">- mature feed mechanism</text>
  <text x="145" y="405" font-family="Arial" font-size="22">- washable wet assembly</text>
  <text x="145" y="450" font-family="Arial" font-size="22">- dry motor boundary</text>
  <text x="145" y="495" font-family="Arial" font-size="22">- real food tests</text>
  <text x="830" y="360" font-family="Arial" font-size="22">- magnetic alignment feel</text>
  <text x="830" y="405" font-family="Arial" font-size="22">- mechanical latch click</text>
  <text x="830" y="450" font-family="Arial" font-size="22">- dummy smart modules</text>
  <text x="830" y="495" font-family="Arial" font-size="22">- crowdfunding demo story</text>
  <line x1="620" y1="415" x2="780" y2="415" stroke="#1f2933" stroke-width="5"/>
  <path d="M780 415 l-18 -12 l0 24 z" fill="#1f2933"/>
  <text x="605" y="710" font-family="Arial" font-size="24" font-weight="700">Integrate only after both gates pass</text>
""",
    )
    write_text(PKG / "drawings" / "DRAWING_INDEX_V0.2.md", """
# AquaClick V0.2 Drawing Index

These are engineering communication drawings, not production drawings, mold drawings or validated CAD outputs.

## Drawing Set

| File | Use |
|---|---|
| `AquaClick_V0.2_GA_Wet_Dry_Section.svg` | Overall wet/dry layout, washable boundary and rough envelope |
| `AquaClick_V0.2_Washable_Food_Path_Detail.svg` | Food-contact washable path: hopper, rotor, liner and bowl |
| `AquaClick_V0.2_Click_Interface_Detail.svg` | Magnetic alignment, mechanical lock and gasket/labyrinth concept |
| `AquaClick_V0.2_Dry_Base_Envelope.svg` | Dry base components: motor, PCB, power, load-cell path and module bay |
| `AquaClick_V0.2_Dual_Prototype_Roadmap.svg` | Split between engineering feeder mule and Click experience model |
| `AquaClick_V0.1C_Eccentric_Hopper_Section.svg` | Earlier V0.1C hopper slope proxy |
| `AquaClick_V0.1C_Washable_Chute_Liner.svg` | Earlier V0.1C chute liner proxy |
| `AquaClick_V0.1C_Assembly_Boundary.svg` | Earlier washable/non-washable boundary sketch |

## OEM Review Questions

1. Which wet/dry boundary is most manufacturable?
2. Can the rotor stay in the washable assembly while the motor remains in the dry base?
3. What latch geometry gives safe retention without making cleaning awkward?
4. What gasket/labyrinth strategy is appropriate before formal IP testing?
5. What prototype process can test the click feel before committing to tooling?
""")


def make_geometry() -> list[dict[str, str | float]]:
    results = []
    for cfg in [HopperConfig("od180", 180.0), HopperConfig("od185", 185.0)]:
        result = make_hopper_stl(cfg)
        results.append({"part": cfg.name, **result})

    for key, spec in ROTOR_VARIANTS.items():
        make_cylinder_stl(PKG / "cad" / "stl" / f"aquaclick_v01c_rotor_{key}_dynamic_envelope.stl", f"rotor_{key}_dynamic_envelope", spec["diameter"] + 2 * spec["clearance"], spec["height"] + 2)
        results.append({"part": f"rotor_{key}_dynamic_envelope", "diameter_mm": spec["diameter"] + 2 * spec["clearance"], "height_mm": spec["height"] + 2, "stl": f"aquaclick_v01c_rotor_{key}_dynamic_envelope.stl"})

    for key, spec in CHANNEL_VARIANTS.items():
        outer_w = spec["inner_w"] + 2 * spec["wall"]
        outer_h = spec["inner_h"] + 2 * spec["wall"]
        make_box_stl(PKG / "cad" / "stl" / f"aquaclick_v01c_chute_liner_{key}.stl", f"chute_liner_{key}", spec["length"], outer_w, outer_h)
        results.append({"part": f"chute_liner_{key}", "inner_w_mm": spec["inner_w"], "inner_h_mm": spec["inner_h"], "length_mm": spec["length"], "stl": f"aquaclick_v01c_chute_liner_{key}.stl"})

    make_cylinder_stl(PKG / "cad" / "stl" / "aquaclick_v01c_floating_coupler_envelope.stl", "floating_coupler_envelope", 34, 42)
    make_box_stl(PKG / "cad" / "stl" / "aquaclick_v01c_dual_side_quick_release_envelope.stl", "dual_side_quick_release_envelope", 126, 28, 34)
    make_box_stl(PKG / "cad" / "stl" / "aquaclick_v01c_center_lever_quick_release_envelope.stl", "center_lever_quick_release_envelope", 92, 38, 44)
    make_box_stl(PKG / "cad" / "stl" / "aquaclick_v01c_load_cell_force_path_envelope.stl", "load_cell_force_path_envelope", 150, 96, 18)
    return results


def pass_section(rotor: dict[str, float], window: dict[str, float], channel: dict[str, float]) -> tuple[float, str]:
    # Conservative digital proxy: the limiting rectangular opening after clearance and draft margin.
    diameter_margin = max(0.0, window["w"] - rotor["clearance"] * 2)
    vertical_margin = max(0.0, min(window["h"], channel["inner_h"]) - 2.0)
    channel_margin = max(0.0, channel["inner_w"] - 2.0)
    min_axis = min(diameter_margin, vertical_margin, channel_margin)
    status = "PASS_15MM_PROXY" if min_axis >= 15.0 else "RISK_LT_15MM"
    return min_axis, status


def write_csvs() -> None:
    rows = []
    for r_key, rotor in ROTOR_VARIANTS.items():
        for w_key, window in WINDOW_VARIANTS.items():
            for c_key, channel in CHANNEL_VARIANTS.items():
                min_axis, status = pass_section(rotor, window, channel)
                rows.append({
                    "rotor": r_key,
                    "window": w_key,
                    "channel": c_key,
                    "minimum_pass_axis_mm_proxy": f"{min_axis:.1f}",
                    "status": status,
                    "note": "Proxy only. Irregular freeze-dried food must be EVT tested.",
                })
    with (PKG / "data" / "rotor_window_channel_clearance_matrix_v01c.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    risks = [
        ["R-P0-001", "Eccentric hopper slope is digital-only", "Food bridging or stagnant zones", "EVT flow test with 5-15 mm kibble and <=20% freeze-dried mix", "Open"],
        ["R-P0-002", "Rigid rotor envelope is not silicone deformation", "False clearance confidence", "Prototype flexible rotor and inspect crumb rate / jam rate", "Open"],
        ["R-P0-003", "Chute liner snap tabs are placeholders", "OEM DFM may require geometry changes", "OEM review with material, draft, tool split and latch force", "Open"],
        ["R-P0-004", "Magnet cannot be safety lock", "Pet/child accidental release", "Mechanical lock retention and misuse test", "Open"],
        ["R-P1-001", "Load cell force path is envelope only", "Weight drift or overload damage", "Bench fixture with 2 kg load-cell and overload stops", "Open"],
    ]
    with (PKG / "data" / "risk_register_v01c.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "risk", "impact", "mitigation", "status"])
        writer.writerows(risks)

    teardown_rows = [
        ["Petlibro Granary / Air class", "mature dry-food dispensing, cleaning, backup power", "rotor geometry, outlet shape, removable tank boundary, bowl interface", "feeding mechanism reference"],
        ["PETKIT Fresh Element / Yumshare direction", "sealed storage, smart sensing, premium pet-tech positioning", "food path separation, hygiene claims, module/camera integration", "hygiene and smart module benchmark"],
        ["Whisker Feeder-Robot class", "premium automatic feeder UX and reliability positioning", "hopper access, programming UX, app promise, service model", "premium UX benchmark"],
        ["Amazon high-volume budget feeder A", "low-cost manufacturable baseline", "motor/gearbox, chute simplicity, BOM reduction, failure complaints", "cost-down benchmark"],
        ["Amazon high-volume budget feeder B", "alternate low-cost dry-food architecture", "anti-jam claims, latch design, cleaning complaints", "failure-mode benchmark"],
    ]
    with (PKG / "data" / "competitor_teardown_matrix_v02.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["target", "why_this_target", "inspect", "decision_use"])
        writer.writerows(teardown_rows)

    mechanism_rows = [
        ["M-FEED-01", "segmented portion wheel", "simple, mature, portion repeatable", "may trap residue, fixed portion granularity", "benchmark as baseline"],
        ["M-FEED-02", "flexible vane rotor", "tolerates irregular kibble better, familiar feeder pattern", "silicone deformation and cleaning residue must be tested", "prototype A/B/C only as adaptation"],
        ["M-FEED-03", "auger assist", "good for metering certain dry foods", "harder to clean, more crevices, higher torque path risk", "secondary only if rotor fails"],
        ["M-FEED-04", "gravity gate / flap", "lowest complexity", "poor small-dose precision and bridging risk", "use only as low-cost reference"],
    ]
    with (PKG / "data" / "mature_feeding_mechanism_scorecard_v02.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "mechanism", "strength", "risk", "aquaclick_decision"])
        writer.writerows(mechanism_rows)


def write_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    sheets = {
        "Inputs": [
            ["Parameter", "Value", "Unit", "Source/Note"],
            ["Food boundary", "dry kibble + limited freeze-dried", "", "Frozen from handoff"],
            ["Particle size", "5-15", "mm", "EVT must verify"],
            ["Freeze-dried ratio", "<=20", "% volume", "Planning input only"],
            ["Usable capacity target", "3.35", "L", "Rechecked in V0.1C proxy"],
            ["Hopper OD variants", "180 / 185", "mm", "EVT comparison"],
        ],
        "Clearance Matrix": [["See CSV", "rotor_window_channel_clearance_matrix_v01c.csv"]],
        "EVT Matrix": [
            ["Test ID", "Purpose", "Sample", "Pass Gate"],
            ["EVT-FLOW-01", "Dose repeatability by food type", "A/B/C rotor x window x channel", "CV and jam rate below agreed threshold"],
            ["EVT-WASH-01", "Wash/rinse/dry cycle", "washable hopper + rotor + liner", "No retained odor, visible residue, or seal damage"],
            ["EVT-LOCK-01", "Quick-release retention and misuse", "dual-side + center-lever", "No accidental release under defined load"],
        ],
        "DFMEA": [
            ["Item", "Failure Mode", "Effect", "Cause", "S", "O", "D", "RPN", "Action"],
            ["Rotor", "Jam with 15 mm irregular food", "Missed feeding", "passage too small / freeze-dried shape", 9, 5, 4, "=E2*F2*G2", "Run clearance + EVT flow matrix"],
            ["Chute liner", "Residue trapped near latch", "Hygiene risk", "dead corner / insufficient radius", 8, 4, 5, "=E3*F3*G3", "OEM DFM + wash residue test"],
            ["Quick release", "Accidental release", "spill / pet access", "magnet-only retention or weak latch", 8, 3, 4, "=E4*F4*G4", "Mechanical lock retention test"],
            ["Load cell", "Overload drift", "wrong dose feedback", "no hard stop", 7, 4, 5, "=E5*F5*G5", "Add overload stop and bench test"],
        ],
        "OEM RFQ Checklist": [
            ["Attachment", "Purpose", "Status"],
            ["V0.1C geometry STL/SVG", "Discuss DFM, tool split, latch geometry", "Ready for discussion, not production data"],
            ["Risk register", "Align engineering unknowns", "Ready"],
            ["EVT matrix", "Quote prototype and test support", "Ready"],
            ["Material candidates", "Confirm food-contact compliance", "OEM/lab required"],
        ],
        "Gate Review": [
            ["Gate", "Requirement", "V0.1C Status"],
            ["Digital geometry generated", "All V0.1C STL/SVG files present", "Pass"],
            ["Clearance proxy", "All combinations checked for 15 mm proxy", "Pass with caveat"],
            ["Physical validation", "Real food / wash / torque / noise / certification", "Not started"],
        ],
    }
    for title, data in sheets.items():
        ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
        for row in data:
            ws.append(row)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="385A64")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 24
    wb.save(PKG / "data" / "AquaClick_EVT_OEM_Control_V0.1C.xlsx")


def write_docs(geometry: list[dict[str, str | float]]) -> None:
    today = date.today().isoformat()
    geometry_lines = "\n".join([f"- `{g.get('part')}`: {', '.join(f'{k}={v}' for k, v in g.items() if k != 'part')}" for g in geometry])
    write_text(PKG / "README.md", f"""
# AquaClick V0.1C Digital Engineering Package

Generated: {today}

This package continues the V0.1B baseline described in `AquaClick_Codex_Project_Handoff_2026-07-16.md`.
The original V0.1B folder and CAD script were not present in the current workspace, so V0.1C is a traceable reconstruction from the handoff record, not a direct edit of missing V0.1B CAD.

## Scope

- Eccentric compound-slope hopper proxies for OD 180 mm and OD 185 mm.
- Washable chute liner proxies with wall thickness, draft, corner-radius assumptions and latch placeholders.
- A/B/C rotor dynamic envelope proxies and rotor/window/channel minimum-passage matrix.
- Floating coupler, quick-release and load-cell force-path envelope models.
- EVT/OEM control workbook, DFMEA seed, risk register and OEM RFQ checklist.

## Important Limit

These files are digital engineering discussion assets. They are not production drawings, mold data, validated STEP geometry, or proof that real food will flow. Food flow, silicone deformation, motor torque, noise, wash reliability, material compliance and OEM DFM still require physical or supplier validation.
""")
    write_text(PKG / "docs" / "RELEASE_NOTES_V0.1C.md", f"""
# AquaClick V0.1C Release Notes

## What Changed From The V0.1B Handoff Baseline

1. Replaced the simplified axisymmetric hopper assumption with an eccentric compound floor proxy that slopes toward an offset rotor inlet.
2. Recomputed usable-capacity proxies for OD 180 and OD 185 variants after reserving volume for rotor/outlet intrusion and safety headspace.
3. Added washable chute liner geometry proxies with wall thickness, draft, corner radius, snap-tab placeholders and finger-access notes.
4. Added A/B/C rotor dynamic envelope STLs and a rotor/window/channel minimum-passage proxy matrix.
5. Added envelope models for floating coupler, dual-side quick release, center-lever quick release and load-cell force path.
6. Added DFMEA seed, risk register, EVT matrix and OEM RFQ checklist.

## V0.1C.1 Clearance Adjustment

After the first V0.1C clearance run, `Rotor C + Window A` measured 14.8 mm in the minimum-passage proxy for both standard and wide channels. `Window A` width has therefore been widened from 18.0 mm to 19.5 mm in this package build. This makes all proxy combinations meet or exceed the 15 mm check, but it does not prove real-food anti-jam performance.

## Files Generated

{geometry_lines}

## Version Caveat

Because the original `AquaClick_V0.1B/` and `build_aquaclick_cad_v01b.py` were not found in the workspace, this package preserves continuity through the handoff document and explicit assumptions. It should be reconciled with the original V0.1B package if that package is later restored.
""")
    write_text(PKG / "reports" / "GEOMETRY_CHECK_REPORT_V0.1C.md", f"""
# Geometry Check Report V0.1C

## Digital Checks Completed

- STL files generated as ASCII mesh assets in millimeter-scale assumptions.
- OD 180 and OD 185 hopper variants include an offset low point near the rotor inlet.
- Rotor A/B/C dynamic envelopes generated as clearance cylinders.
- Chute liner standard/wide variants generated as outer-envelope proxies.
- Clearance matrix generated for all rotor x window x channel combinations.

## Capacity Proxy

The hopper capacity is computed as a conservative proxy, not a fluid/particle fill simulation. It deducts approximate rotor/outlet displacement and safety headspace. Real fill behavior with kibble and freeze-dried pieces must be tested.

## Known Digital Limitations

- No OpenCascade/CadQuery kernel is available in this environment, so STEP export and B-rep validation are not included.
- Rounded corners, snap tabs, latch faces and flexible silicone vanes are represented as placeholders or envelopes.
- Dynamic rotor simulation is a swept-volume proxy, not a contact/friction simulation.
""")
    write_text(PKG / "docs" / "OEM_RFI_RFQ_ATTACHMENT_GUIDE_V0.1C.md", """
# OEM RFI/RFQ Attachment Guide V0.1C

## Send To OEM

- `README.md`
- `docs/RELEASE_NOTES_V0.1C.md`
- `reports/GEOMETRY_CHECK_REPORT_V0.1C.md`
- `data/AquaClick_EVT_OEM_Control_V0.1C.xlsx`
- `data/risk_register_v01c.csv`
- `data/rotor_window_channel_clearance_matrix_v01c.csv`
- `drawings/*.svg`
- `cad/stl/*.stl`

## Ask OEM To Confirm

1. Food-contact material options for transparent hopper and washable liner.
2. Injection molding feasibility for liner wall thickness, draft, snap tabs and rounded corners.
3. Tooling and part-cost estimate for 180 mm vs 185 mm hopper.
4. Prototype method for flexible rotor variants.
5. Latch retention force, release force, spring selection and expected cycle life.
6. Load-cell mounting strategy, overload stop and liquid ingress boundary.

## Do Not Claim Yet

- Dishwasher safe.
- IPX7 or any waterproof rating.
- Anti-jam performance with all kibble.
- Multi-cat accurate feeding using vision alone.
- FDA/EU food-contact compliance before material and migration confirmation.
""")
    write_text(PKG / "docs" / "PHYSICAL_VALIDATION_REQUIRED_V0.1C.md", """
# Items That Still Require Physical / Lab / OEM Validation

- Real 5-15 mm kibble and freeze-dried flow, bridging, jamming and dose repeatability.
- Flexible silicone rotor deformation, fatigue, crumb rate and cleaning residue.
- Motor torque, stall current, temperature rise, noise and life.
- Quick-release retention, child/pet misuse, drop survival and cycle life.
- Wash, soak, dry, odor, mold risk and seal reliability after repeated cleaning.
- Food-contact material compliance and migration testing for US/EU markets.
- OEM DFM, tool split, draft, tolerance, official BOM, MOQ, lead time and quote.
- Camera privacy and wireless compliance for FCC/CE/UKCA if smart variants proceed.
""")
    write_text(PKG / "docs" / "EVT_SAMPLE_GATE_V0.1C.md", """
# AquaClick V0.1C EVT Sample Gate

## Gate Intent

This gate defines what must be true before AquaClick V0.1C geometry is treated as ready for EVT sample fabrication discussion. Passing this gate does not mean the product is validated; it only means the digital package is coherent enough to ask OEM/prototype partners for cost, DFM feedback and sample planning.

## Digital Gate Criteria

| ID | Criterion | V0.1C.1 Status |
|---|---|---|
| G-DIG-01 | Food boundary fixed as dry kibble + limited freeze-dried, 5-15 mm, <=20% freeze-dried volume for EVT planning | Pass |
| G-DIG-02 | OD 180 and OD 185 eccentric hopper variants generated | Pass |
| G-DIG-03 | Rotor A/B/C dynamic envelopes generated | Pass |
| G-DIG-04 | Rotor x window x channel minimum-passage proxy checks >= 15 mm | Pass after Window A widened to 19.5 mm |
| G-DIG-05 | Washable liner has wall/draft/radius/latch assumptions documented | Pass as proxy |
| G-DIG-06 | Washable and dry electrical boundaries documented | Pass |
| G-DIG-07 | DFMEA seed and risk register available | Pass |
| G-DIG-08 | Physical validation items explicitly separated from digital checks | Pass |

## OEM Must Answer Before EVT Tooling Or Prototype Spend

1. Whether OD 180 mm can hold target usable capacity after real ribs, draft, outlet intrusion and molded detail are added, or whether OD 185 mm should become EVT default.
2. Whether the widened Window A causes unacceptable rotor leakage, dose scatter, wall weakness or tool complexity.
3. Whether the washable chute liner snap-fit can survive repeated removal without residue traps.
4. Whether the quick-release mechanism should continue with dual-side, center-lever, or both for EVT.
5. Which material stack is realistic for transparent food-contact hopper, liner, rotor and seal.
6. Which prototype process can approximate flexible silicone rotor behavior enough for EVT learning.

## Do Not Claim At This Gate

- No jam / anti-jam.
- Dishwasher safe.
- IPX rating.
- Accurate dosing across all kibble.
- Food-contact compliance.
- Production-ready CAD, mold-ready CAD or validated STEP.

## Recommended EVT Sample Matrix

| Axis | Options |
|---|---|
| Hopper OD | 180 mm, 185 mm |
| Rotor | A, B, C |
| Window | A widened, B, C |
| Chute | standard, wide |
| Food | small dry kibble, mixed dry kibble, large 15 mm kibble, 20% freeze-dried mix |
| Quick release | dual-side, center-lever |

Minimum practical first pass: choose OD 185, Rotor A/B/C, Window A/B/C, standard/wide chute and four food mixes. Keep OD 180 as a capacity-risk comparison sample if budget allows.
""")
    write_text(PKG / "docs" / "执行摘要_V0.1C.md", """
# AquaClick V0.1C 执行摘要

## 接手状态

已读取 `AquaClick_Codex_Project_Handoff_2026-07-16.md`。当前工作区没有找到原交接文件中提到的 `AquaClick_V0.1B/`、`build_aquaclick_cad_v01b.py` 和 V0.1B ZIP，因此本 V0.1C 包是基于交接文件冻结基线的可追溯重建，不声称直接继承或修改了缺失的 V0.1B CAD。

## 本轮推进

- 建立 OD 180 / OD 185 两个偏心复合坡面食物仓代理模型，目标是让粮食自然汇聚到偏置转子入口。
- 建立标准 / 加宽可拆洗出粮通道内衬代理模型，包含壁厚、拔模、圆角、卡扣占位与手指操作空间说明。
- 建立 A/B/C 转子动态旋转包络，并输出转子 x 窗口 x 通道的最小通过截面代理矩阵。
- 建立浮动联轴器、双侧快拆、中央杠杆快拆和称重传力路径的空间包络。
- 补齐 EVT/OEM 控制工作簿、DFMEA 初版、风险清单、OEM RFI/RFQ 附件指南和数字几何检查报告。

## 数字检查结论

- 首轮 V0.1C 中，18 个转子/窗口/通道组合有 16 个通过 15 mm 代理截面检查，2 个风险项集中在 `Rotor C + Window A`。
- V0.1C.1 已将 `Window A` 宽度从 18.0 mm 调整为 19.5 mm；重新生成后所有组合通过 15 mm 代理截面检查。
- 代理截面通过不能用于宣称“不卡粮”，真实干粮和冻干仍需 EVT 实测。
- 当前模型可用于工程讨论、EVT 准备和 OEM DFM 沟通，但不能作为生产图、开模数据或性能结论。

## 仍需实测 / OEM 确认

真实干粮和冻干的流动与卡粮、柔性硅胶叶片变形、电机扭矩与噪声、快拆寿命与误触、反复清洗后的密封可靠性、食品接触材料法规、OEM 模具和正式报价，都不能由本数字包替代。
""")
    if HANDOFF.exists():
        shutil.copy2(HANDOFF, PKG / "docs" / HANDOFF.name)

    write_text(PKG / "docs" / "PRODUCT_STRATEGY_SHIFT_V0.2.md", """
# AquaClick V0.2 Product Strategy Shift

## Decision

Do not make the feeding rotor the core invention. Treat food dispensing as a mature mechanism to adapt, verify and package cleanly. AquaClick's real research focus moves to:

1. Circuit-free food-contact assembly.
2. Fully washable wet-zone architecture.
3. Magnetic alignment plus mechanical lock connection.
4. Clean power/drive boundary between washable and dry zones.
5. Click module interface and extension ecosystem.

## Why

Automatic feeders already compete on scheduled feeding, portioning, backup power, camera monitoring, app control and basic anti-jam performance. A new rotor is unlikely to be the reason customers understand or fund AquaClick. A visibly removable, washable food path and satisfying Click expansion system is easier to defend, demonstrate and market.

## New P0 Priorities

| Rank | Workstream | Goal |
|---|---|---|
| P0-1 | Wet-zone architecture | Food-contact path contains no circuit, motor, battery, PCB, camera or permanent wiring |
| P0-2 | Washable assembly | Hopper, rotor, chute liner and bowl can be removed as a cleanable system |
| P0-3 | Dry-zone boundary | Motor, load cell, PCB, camera and power remain isolated from rinse/soak exposure |
| P0-4 | Magnetic + mechanical connection | Magnets provide alignment and feel; mechanical locks provide safety retention |
| P0-5 | Mature feeding adaptation | Benchmark and adapt proven dispenser architectures instead of inventing from scratch |
| P0-6 | Click interface | Define module envelope, lock, contacts, ID and safety limits |

## What Gets Deprioritized

- Novel rotor patents unless testing proves a clear need.
- App feature depth beyond reliable scheduling and clear maintenance feedback.
- Advanced AI multi-cat claims before controlled feeding hardware is solved.
- Decorative modules before the base Click interface is mechanically reliable.
""")

    write_text(PKG / "docs" / "WET_DRY_ARCHITECTURE_V0.2.md", """
# AquaClick V0.2 Wet / Dry Architecture

## Architecture Principle

Everything that touches food or receives food dust should be removable and washable. Everything electrical stays in the dry base or sealed expansion modules.

## Wet Zone

- Transparent food hopper.
- Hopper lid seal and desiccant holder if washable boundary permits.
- Removable rotor / dispensing wheel.
- Removable chute liner.
- Bowl or bowl tray.
- Non-electrical mechanical coupler insert if food-contact-safe.

Wet zone must not contain:

- PCB.
- Battery.
- Motor.
- Permanent wire harness.
- Camera.
- Load-cell electronics.
- Speaker or microphone.
- Non-sealed pogo/contact board.

## Dry Zone

- Motor and gearbox.
- Main PCB.
- Power input and backup battery.
- Load-cell electronics.
- Camera / AI module.
- Speaker, microphone, display or status LED.
- Pogo/contact controller for dry expansion modules.

## Boundary Requirements

| Boundary | Requirement |
|---|---|
| Mechanical drive | Torque transfers through a detachable coupler; wet part can be washed separately |
| Water/dust | Labyrinth or gasket boundary prevents rinse water and food dust from entering motor cavity |
| User action | User can remove wet zone without tools |
| Safety | Unit cannot run when wet assembly is absent or unlocked |
| Cleaning | No hidden dead corner in normal food-contact flow path |

## Open Engineering Questions

1. Can the rotor stay entirely in the wet zone while the motor remains in the dry zone?
2. Should the motor drive through a vertical shaft, horizontal shaft or keyed floating coupler?
3. Can the load-cell system remain under the bowl without compromising washability?
4. Is a dishwasher-safe claim strategically worth the validation burden, or should V1 claim rinse/soak/hand-wash only?
""")

    write_text(PKG / "docs" / "CLICK_INTERFACE_SPEC_V0.1.md", """
# AquaClick Click Interface Spec V0.1

## Purpose

Define a repeatable physical and electrical interface for AquaClick modules. The first use is the washable food assembly connection. Later uses include camera, AI, expression display and co-branded modules.

## Interface Layers

| Layer | Role | V0.1 Direction |
|---|---|---|
| Alignment | Fast user positioning and premium feel | Magnet pairs plus lead-in geometry |
| Retention | Prevent lift-off, pet misuse and vibration release | Mechanical hooks / latch, not magnet-only |
| Sealing | Separate washable wet zone and dry electronics | Gasket, labyrinth, raised splash boundary |
| Drive | Transfer torque to removable rotor if required | Floating keyed coupler with axial/radial tolerance |
| Electrical | Optional for dry modules only | Pogo pins or sealed connector, no wet food path contacts |
| Identity | Module recognition | resistor ID / 1-wire / low-speed handshake, later decision |

## Mechanical Envelope Draft

- Primary module insertion: vertical drop-in with front tactile Click.
- Alignment tolerance target: +/- 1.0 mm before magnetic capture.
- Retention target: enough to resist pet pawing and accidental lift; exact force must be bench tested.
- Release action: deliberate two-step action for wet food assembly; one-handed acceptable only if mis-release is controlled.
- Visible feedback: physical seated line plus app/LED confirmation if dry electronics can detect lock state.

## Electrical Guardrails

- Food-contact wet assembly should be passive by default.
- If a module needs power, it must be outside the washable food path or independently sealed.
- No mandatory cloud subscription for core feeding schedule.
- Module ID must not allow unapproved high-power modules without safety negotiation.

## Module Classes

| Class | Examples | Allowed In V1 |
|---|---|---|
| Passive wet module | hopper, chute, bowl trim, food-safe collab insert | Yes |
| Dry smart module | camera, expression display, status light | Yes after connector validation |
| Regulated food-contact module | treat insert, special food insert | Later; needs material and liability review |
| Third-party powered module | partner hardware | Later; requires certification program |

## Next Spec Decisions

1. Magnet grade, count, placement and shielding.
2. Latch geometry and release action.
3. Coupler torque target and tolerance stack.
4. Seated/locked detection method.
5. Connector pinout for dry modules.
""")

    write_text(PKG / "docs" / "COMPETITOR_TEARDOWN_PLAN_V0.2.md", """
# Competitor Teardown Plan V0.2

## Goal

Use mature feeders to avoid reinventing the dispensing mechanism, while identifying the exact gaps AquaClick can own: washable wet zone, dry electronics separation and Click modularity.

## Targets

See `data/competitor_teardown_matrix_v02.csv`.

## Teardown Checklist

| Area | What To Measure |
|---|---|
| Hopper | capacity, wall thickness, outlet angle, dead corners, lid seal, desiccant path |
| Dispensing | rotor/wheel/auger geometry, outlet size, anti-bridge features, motor position |
| Cleaning | removable parts, dishwasher/hand-wash claims, hidden residue traps |
| Dry/wet boundary | where electronics sit, how dust/water are isolated |
| Reliability | backup power, lockouts, jam detection, empty-food detection |
| Cost | part count, fasteners, snap fits, molded complexity |
| UX | refill, cleaning, setup, portion setting, pet-proofing |

## Output Template

For each product, produce:

1. Photos of all removed assemblies.
2. Food-path sketch.
3. Motor-to-rotor drive sketch.
4. Cleaning boundary sketch.
5. List of parts AquaClick should copy as mature pattern.
6. List of parts AquaClick should deliberately improve.

## Success Definition

After teardown, AquaClick should choose one mature dispensing reference architecture for EVT, with one backup architecture. The product should not continue with three equally weighted rotor inventions unless teardown/testing proves it is necessary.
""")

    write_text(PKG / "docs" / "FEEDING_MECHANISM_ADAPTATION_PLAN_V0.2.md", """
# Mature Feeding Mechanism Adaptation Plan V0.2

## Decision Frame

The feeder mechanism is a reliability subsystem, not the brand-defining subsystem. The goal is not novelty; the goal is predictable dry-food dispensing inside AquaClick's washable wet-zone architecture.

## Candidate Mechanisms

See `data/mature_feeding_mechanism_scorecard_v02.csv`.

## Recommended Path

1. Benchmark segmented portion wheel and flexible vane rotor first.
2. Use auger assist only if rotor/wheel approaches fail with freeze-dried mixes.
3. Treat gravity gate as low-cost reference, not primary AquaClick architecture.

## EVT Metrics

| Metric | Why It Matters |
|---|---|
| Jam rate | Core reliability |
| Portion repeatability | Feeding trust |
| Crumb rate | Food quality and hygiene |
| Residue after rinse | AquaClick hygiene promise |
| Noise | Home/pet acceptance |
| Motor torque margin | Cost and reliability |
| User cleaning time | Product differentiation |

## Design Constraint

The chosen mechanism must support a removable wet-side food path. If a proven mechanism cannot be made washable without bringing electronics/motor into the food-contact assembly, it is not suitable for AquaClick's positioning.
""")

    write_text(PKG / "docs" / "DUAL_PROTOTYPE_PLAN_V0.2.md", """
# Dual Prototype Plan V0.2

## Why Two Prototypes

One prototype cannot efficiently answer both engineering reliability and crowdfunding desirability. Split the work.

## Prototype A: Engineering Feeder Mule

Purpose: prove the washable wet-zone feeder architecture works with mature dispensing mechanisms.

Includes:

- OD 185 hopper default, OD 180 optional comparison.
- Mature segmented wheel or flexible rotor reference mechanism.
- Removable chute liner.
- Dry base with motor isolated from wash zone.
- Basic lockout sensor if assembly is missing or unlocked.

Does not need:

- Final exterior design.
- Expression display.
- Full app.
- Partner modules.

## Prototype B: Click Experience Model

Purpose: prove the product is understandable, delightful and crowdfundable.

Includes:

- Magnetic alignment feel.
- Mechanical latch sound/action.
- Camera/display dummy modules.
- Co-branded passive module mockups.
- Clear washable assembly removal demo.

Does not need:

- Real food dispensing.
- Full electronics.
- Final material stack.

## Gate

Proceed to integrated EVT only after Prototype A proves basic food-path reliability and Prototype B proves the Click story is visibly compelling.
""")

    write_text(PKG / "docs" / "RND_RISK_REORDER_V0.2.md", """
# AquaClick R&D Risk Reorder V0.2

## Old Risk Order

1. Rotor geometry.
2. Channel clearance.
3. Hopper slope.
4. Quick release.
5. Coupler.
6. Load cell.

## New Risk Order

| Rank | Risk | Reason |
|---|---|---|
| 1 | Wet-zone contains no electronics but still dispenses reliably | Core product claim depends on it |
| 2 | Wet/dry boundary survives real cleaning behavior | Failure damages electronics and trust |
| 3 | Magnetic alignment plus mechanical lock feels good and is safe | Brand experience and safety both depend on it |
| 4 | Mature dispensing mechanism can be adapted into washable assembly | Avoids over-inventing while keeping differentiation |
| 5 | Click interface can support future modules without unsafe complexity | Long-term ecosystem depends on early interface choices |
| 6 | Cost and DFM remain viable | Premium story fails if BOM explodes |
| 7 | Smart features add value without privacy/subscription backlash | Important, but secondary to physical promise |

## Immediate Kill Criteria

- If wet assembly cannot be removed and rinsed without exposing electronics, AquaClick loses its core differentiation.
- If mechanical lock requires awkward two-handed force or can be released by pet interaction, Click interface must be redesigned.
- If mature feeding mechanisms cannot be adapted without severe residue traps, reposition from "fully washable food path" to a narrower claim before crowdfunding.
""")

    write_text(PKG / "docs" / "NEXT_ACTION_INDEX_V0.2.md", """
# AquaClick Next Action Index V0.2

## Read First

1. `PRODUCT_STRATEGY_SHIFT_V0.2.md` - confirms the pivot: mature feeding mechanism adaptation, AquaClick owns washable wet-zone and Click modularity.
2. `WET_DRY_ARCHITECTURE_V0.2.md` - defines what belongs in washable wet zone versus dry electronics zone.
3. `CLICK_INTERFACE_SPEC_V0.1.md` - first-pass rules for magnetic alignment, mechanical lock, sealing, drive and module identity.

## Then Execute

1. Run the competitor teardown work in `COMPETITOR_TEARDOWN_PLAN_V0.2.md` using `data/competitor_teardown_matrix_v02.csv`.
2. Pick the first dispensing reference architecture using `FEEDING_MECHANISM_ADAPTATION_PLAN_V0.2.md` and `data/mature_feeding_mechanism_scorecard_v02.csv`.
3. Build two separate prototypes from `DUAL_PROTOTYPE_PLAN_V0.2.md`: Engineering Feeder Mule and Click Experience Model.
4. Use `RND_RISK_REORDER_V0.2.md` to keep engineering attention on wet/dry separation, lock safety and module interface instead of drifting back into novel rotor invention.

## Next Concrete Purchase / Build Decisions

- Choose 3-5 competitor feeders to buy or obtain for teardown.
- Decide whether OD 185 mm becomes the default engineering mule hopper.
- Choose first-pass mature dispensing mechanism: segmented wheel or flexible vane rotor.
- Decide whether Click wet assembly release should be dual-side, center-lever or both in first prototype.
- Decide whether dry smart modules use electrical contacts in V1 or remain passive/dummy for crowdfunding prototype.

## Definition Of Done For The Next Sprint

- At least three competitor feeders torn down and photographed.
- One mature dispensing architecture selected as EVT primary, one as backup.
- Wet-zone/dry-zone boundary sketch reviewed with OEM or prototype shop.
- Click latch and magnet placement sketched with target retention/release forces.
- Prototype A and Prototype B build scopes priced separately.
""")


def write_parameter_table(geometry: list[dict[str, str | float]]) -> None:
    with (PKG / "data" / "geometry_summary_v01c.csv").open("w", newline="", encoding="utf-8") as f:
        keys = sorted({key for row in geometry for key in row.keys()})
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(geometry)


def make_zip() -> Path:
    OUT.mkdir(exist_ok=True)
    zip_path = OUT / "AquaClick_Digital_Engineering_Package_V0.1C.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for file in PKG.rglob("*"):
            if file.is_file():
                z.write(file, file.relative_to(ROOT))
    return zip_path


def main() -> None:
    reset_dirs()
    geometry = make_geometry()
    make_drawings()
    write_csvs()
    write_workbook()
    write_parameter_table(geometry)
    write_docs(geometry)
    zip_path = make_zip()
    print(f"generated={PKG}")
    print(f"zip={zip_path}")


if __name__ == "__main__":
    main()
